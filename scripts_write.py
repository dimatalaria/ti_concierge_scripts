import openpyxl

keynect = openpyxl.load_workbook("keystrom.xlsx")

sheet = keynect.active

i = 1

num = 1

while i <= 100:
    sheet.cell(row=i, column=2, value=num)
    sheet.cell(row=i+1, column=2, value=num)
    i += 2
    num += 1

keynect.save("keystrom.xlsx")
