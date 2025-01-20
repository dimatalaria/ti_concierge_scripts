import openpyxl

keynect = openpyxl.load_workbook("losika53.xlsx")

sheet = keynect.active

value_key = sheet["A1"].value

cell_key = sheet.cell(row=1, column=1).value
sps = []
for row in sheet.iter_rows(min_row=1, max_row=606, values_only=True):
    num = row[0]
    hex_number = f"{num:08X}"
    bytes_list = [hex_number[i:i + 2] for i in range(0, len(hex_number), 2)]
    reversed_bytes = ''.join(bytes_list[::-1]).upper()
    json = {"CardNo":reversed_bytes, "UserID":row[1], "UserName":"53", "VTOPosition":"", "CardType":0, "CardStatus":0}
    sps.append(json)


file = open("key.txt", "w")
file.write(str(sps))
file.close()
print(sps)